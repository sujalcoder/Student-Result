from flask import Blueprint, render_template, request , redirect, url_for
from database import teachers, marks ,students, admins
from pymongo.errors import DuplicateKeyError
import csv
from io import TextIOWrapper
admin_bp = Blueprint("admin", __name__)


# ADMIN LOGIN ROUTE
@admin_bp.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # TEMP TEST LOGIN
        if username == "admin" and password == "adminpass":
            return render_template('admin_dashboard.html')
        else:
            return render_template(
                'admin_login.html',
                error='Invalid username or password'
            )

    return render_template('admin_login.html')


# ADMIN DASHBOARD ROUTE
@admin_bp.route('/admin/dashboard', methods=['GET', 'POST'])
def admin_dashboard():
    return render_template('admin_dashboard.html')

# ADD TEACHER ROUTE
@admin_bp.route('/admin/add-teacher', methods=['GET', 'POST'])
def add_teacher():
    if request.method == 'POST':
        teacher_name = request.form['teacher_name']
        gender = request.form['gender']
        section = request.form.getlist('section[]')
        rights = request.form['rights']
        username = request.form['username']
        password = request.form['password']

        # 🔍 Check duplicate username
        existing_teacher = teachers.find_one({'username': username})
        if existing_teacher:
            return render_template(
                'add_teacher.html',     
                error='Teacher username already exists'
            )

        # ✅ Insert teacher
        teachers.insert_one({
            'name': teacher_name,
            'gender': gender,
            'section': section,
            'rights': rights,
            'username': username,
            'password': password   # hashing later
        })

        # 🔁 Redirect to avoid duplicate submit
        return redirect(url_for('admin.manage_teacher'))

    return render_template('add_teacher.html')


# MANAGE TEACHER ROUTE
@admin_bp.route('/admin/manage-teacher', methods=['GET', 'POST'])
def manage_teacher():  
    all_teachers = list(teachers.find())
    return render_template('manage_teacher.html', teachers=all_teachers) 

# UPDATE TEACHER ROUTE
@admin_bp.route('/admin/update-teacher', methods=['POST'])
def update_teacher():
    print("UPDATE ROUTE HIT")
    print(request.form)

    username = request.form['username']
    name = request.form['name']
    password = request.form['password']
    section = request.form.getlist('section[]')
    rights = request.form.get('rights')
    gender = request.form.get('gender')

    update_data = {
        'name': name,
        'section': section,
        'rights': rights,
        'gender': gender
    }

    if password:
        update_data['password'] = password

    result = teachers.update_one(
    {'username': username},
    {'$set': update_data}
)

    print("MATCHED:", result.matched_count)
    print("MODIFIED:", result.modified_count)


    return redirect(url_for('admin.manage_teacher'))

# ADD STUDENT ROUTE
@admin_bp.route('/admin/add-student', methods=['GET', 'POST'])
def add_student():
    if request.method == 'POST':
        student_name = request.form['student_name']
        roll_no = request.form['roll_no']
        section = request.form['section']
        existing_student=students.find_one({'roll_no': roll_no})
        if existing_student:
            return render_template(
                'add_student.html',
                error='Roll No already exists'
            )
        else:

         # ✅ Insert student
            try:
                students.insert_one({
                   'name': student_name,
                   'roll_no': roll_no,
                   'section': section
            })

            # ✅ Redirect after POST (VERY IMPORTANT)
                return redirect(url_for('admin.manage_student'))

            except DuplicateKeyError:
                return render_template(
                  'add_student.html',
                   error='Roll No already exists'
            )

            except Exception as e:
                return render_template(
                  'add_student.html',
                   error='Something went wrong'
            )

    return render_template('add_student.html')

# MANAGE STUDENT ROUTE
@admin_bp.route('/admin/manage-student', methods=['GET', 'POST']) 
 
def manage_student():  
    all_students = students.find()
    return render_template('manage_student.html', students=all_students)

#DELETE STUDENT ROUTE

@admin_bp.route('/admin/delete-student', methods=['POST'])
def delete_student():
    from bson import ObjectId

    student_id = request.form['student_id']

    # delete student
    students.delete_one({'_id': ObjectId(student_id)})

    # also delete marks of that student
    marks.delete_many({'student_id': ObjectId(student_id)})

    return redirect(url_for('admin.manage_student'))

# UPDATE STUDENT ROUTE
@admin_bp.route('/admin/update-student', methods=['POST'])
def update_student():
    from bson import ObjectId

    student_id = request.form['student_id']
    name = request.form['name']
    section = request.form['section']

    students.update_one(
        {'_id': ObjectId(student_id)},
        {'$set': {
            'name': name.strip(),
            'section': section
        }}
    )

    return redirect(url_for('admin.manage_student'))


#delete teacher route
@admin_bp.route('/admin/delete-teacher', methods=['POST'])
def delete_teacher():
    username = request.form['username']
    teachers.delete_one({'username': username})
    return redirect(url_for('admin.manage_teacher'))



from openpyxl import load_workbook

@admin_bp.route('/admin/upload-students', methods=['POST'])
def upload_students():
    file = request.files['file']

    if not file:
        return redirect(url_for('admin.add_student'))

    wb = load_workbook(file)
    sheet = wb.active

    added = 0
    skipped = 0

    # Excel rows read (skip header row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        roll_no, name, section = row

        if not roll_no or not name or not section:
            continue

        # Check duplicate roll no
        if students.find_one({'roll_no': str(roll_no)}):
            skipped += 1
            continue

        students.insert_one({
            'roll_no': str(roll_no),
            'name': name.strip(),
            'section': section
        })
        added += 1

    return redirect(url_for('admin.manage_student'))
